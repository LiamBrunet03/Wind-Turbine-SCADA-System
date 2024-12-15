import tkinter as tk
from tkinter import font
import serial
import threading
import time
from tkinter import messagebox
import datetime
import openpyxl
from openpyxl.chart import LineChart, Reference
import re

window = tk.Tk()
window.title("Servo Angle Control")
window.geometry("1200x700")


current_datetime = datetime.datetime.now().strftime("%B_%d_%Y_%I_%M%p")# gets the current date and time for the title of the log file
        

# variables
total_power = 0
power_value= 0
seconds_counter = 1  # Starts at 1 second for the loging part
servo_angle = 90  # Servo starts at the middle (90 degrees)
is_manual_mode = True  # Starts in automatic mode
is_stopped = False  # Initial state for the stop button
#-------------------------------------------------------------------------------------------------------------------------------------------------

# serial communication
try:
    ser = serial.Serial('COM11', 9600, timeout=1)
except serial.SerialException: #debug 
    print("Could not open serial port. Check your port connection.")
    ser = None # serial value starts at none

#-------------------------------------------------------------------------------------------------------------------------------------------------
#                                                                   Login

def show_main_screen(is_logged_in):# Function to show the main screen after login
   
    login_frame.pack_forget() # delets the login screen

    
    if is_logged_in:# if logged in succesful then show the main screen
        control_frame.pack(fill="both", expand=True)
    else:
        guest_frame.pack(fill="both", expand=True)


def login():# Login function
    username = username_entry.get() 
    password = password_entry.get()

    if username == "liam" and password == "liam":# if the values of username and password match liam then show the main screen
        messagebox.showinfo("Login Successful", "Welcome, Liam!")
        show_main_screen(True)  # open main screen
    else:
        messagebox.showerror("Login Failed", "Incorrect username or password!")
        show_main_screen(False)   # does not open main screen but shows the message "Login Failed", "Incorrect username or password!"

# if the guest button is pressed then this Function that removes all the buttons from the main window will be executed.
def guest_login():
    """Handle guest login."""
    messagebox.showinfo("Guest Login", "Welcome, Guest!")#greetings for the guest
    show_main_screen(True)  # Show the main screen
    automatic_button.grid_remove()# Removes all the buttons so that the guest user can only see the data 
    manual_button.grid_remove()
    up_button.grid_remove()
    up_arrow_label.grid_remove()
    down_arrow_label.grid_remove()
    down_button.grid_remove()
    stop_button.grid_remove()




# Define frames for login and main screens
login_frame = tk.Frame(window)
login_frame.pack(fill="both", expand=True)

control_frame = tk.Frame(window)
guest_frame = tk.Frame(window)

# Creates the login screen labels and buttons
username_label = tk.Label(login_frame, text="Username:", font=("Arial", 12))
username_label.grid(row=9, column=0, columnspan=5, sticky="n")
username_entry = tk.Entry(login_frame, font=("Arial", 12))
username_entry.grid(row=8, column=0, columnspan=5, sticky="n")

password_label = tk.Label(login_frame, text="Password:", font=("Arial", 12))
password_label.grid(row=7, column=0, columnspan=5, sticky="n")
password_entry = tk.Entry(login_frame, font=("Arial", 12), show="*")
password_entry.grid(row=6, column=0, columnspan=5, sticky="n")

# Login button
login_button = tk.Button(login_frame, text="Login", font=("Arial", 14), command=login)
login_button.grid(row=5, column=0, columnspan=5, sticky="n")

# Guest button
guest_button = tk.Button(login_frame, text="Sign in as Guest", font=("Arial", 14), command=guest_login)
guest_button.grid(row=4, column=0, columnspan=5, sticky="n")

#-------------------------------------------------------------------------------------------------------------------------------------------------
#                                                                   Skips login
# Skip the login screen by simulating a successful login if the code under it is not commented out.
  
#show_main_screen(True)
#-------------------------------------------------------------------------------------------------------------------------------------------------
#                                                                   python and arduino comms
# Function to update the angle label
def update_angle_label():
    angle_label.config(text=f"Angle: {servo_angle}°")

# Function to send angle to Arduino
def send_angle_to_arduino(angle):
    global servo_angle
    try:
        ser.write(f'{servo_angle}\n'.encode())  # Send the updated angle over serial
        ser.flush()
        print(f"Sent angle {servo_angle} to Arduino.")
    except Exception as e:
        print("Error sending data:", e)

# Function to toggle the stop button state
def toggle_stop():
    global is_stopped
    is_stopped = not is_stopped
    if is_stopped: #if the on off putton is turned on turn the label to green and display on
        stop_button.config(text="On", bg="green")
        if ser:
            ser.write("STOP_ON\n".encode())  # Send "STOP_ON" to Arduino over serial
    else:
        stop_button.config(text="Off", bg="red")#else the on off putton is turned off and  the label will be to red  and display off
        if ser:
            ser.write("STOP_OFF\n".encode())  # Send "STOP_OFF" to Arduino over serial

# Function to increase angle but only if manual mode is on.
def increase_angle():
    global servo_angle
    if is_manual_mode:
        servo_angle += 10
        ser.write(f'servo {servo_angle},\n'.encode())  # Sends "servo <angle>," to arduino
        update_angle_label()#updates the servo label with the new servo angle

# Function to decrease angle but only if automatic mode is on.
def decrease_angle():
    global servo_angle
    if is_manual_mode:
        servo_angle -= 10
        ser.write(f'servo {servo_angle},\n'.encode())  # Sends "servo <angle>," to arduino
        update_angle_label()#updates the servo label with the new servo angle
        
is_automatic_mode = False# automatic starts false in the hmi


is_manual_mode = True # manual starts true in the hmi

# Functions to toggle between manual and automatic modes
def toggle_automatic_mode():
    
    
    global is_manual_mode,is_automatic_mode
    is_automatic_mode = True # change automatic_mode to true
    manual_button.config(bg="red")
    automatic_button.config(bg="green")
    ser.write(b'servo 90,\n')  # Sends "servo 90," to arduino
    ser.write(b'Automatic \n')
    print("set to automatic")  
    is_manual_mode = False # change maual mode to false




def toggle_manual_mode():
    global is_manual_mode,is_automatic_mode
    is_manual_mode = True # change maual mode to true
    manual_button.config(bg="green")
    automatic_button.config(bg="red")
    ser.write(b'Manual\n')
    print("set to manual ")  
    is_automatic_mode = False # change automatic_mode to false


    


# Function to continuously read data from the arduino every 0.01 seconds
def read_serial_data():
    global servo_angle, turbine_temp_value, outside_temp_value, turbine_voltage_value, turbine_current_value, power_value, total_power, RPM # gets all these variables from the code as a reference
    
    while True:
        if ser and ser.in_waiting > 0:# if their is stuff to read in the serial 
            try:
                line = ser.readline().decode('utf-8').strip()  # Read data from the serial port and converts the byte data into a string so that the if statments can filter keywords

                # Check if the line contains turbine temperature data
                match_temp = re.search(r"Turbine Temperature:\s*(\d+)", line)# checks to see if Turbine Temperature: is found and if so to capture the number.
                if match_temp:#if match was found 
                    turbine_temp_value = match_temp.group(1)  # Extract temperature value from the capture
                    update_label(turbine_temp_value_label, turbine_temp_value)#updates the turbine temp label with the value found.

                # Check if the line contains servo data
                match_servo = re.search(r"Servo:\s*(\d+),", line)
                if match_servo:
                    servo_angle = int(match_servo.group(1))  # Extract servo angle value
                    update_angle_label()  # Update the angle display

                # Check if the line contains outside temperature data
                match_outside_temp = re.search(r"Outside Temperature:\s*(\d+)", line)
                if match_outside_temp:
                    outside_temp_value = match_outside_temp.group(1)  # Extract outside temperature value
                    update_label(outside_temp_value_label, outside_temp_value)
                
                match_outside_temp = re.search(r"RPM:\s*(\d+)", line)
                if match_outside_temp:
                    RPM_value = int(match_outside_temp.group(1))  # Convert the extracted value to an integer
                    if RPM_value > 100:
                        update_label(RPM_value_label, RPM_value)
                    else:
                        update_label(RPM_value_label, 0)
                    

                # Check if the line contains turbine voltage data
                match_voltage = re.search(r"Turbine Voltage:\s*(\d+)", line)
                if match_voltage:
                    turbine_voltage_value = int(match_voltage.group(1)) # Extract turbine voltage value
                    update_label(turbine_voltage_value_label, turbine_voltage_value)
                    
                    if turbine_voltage_value > 4 and is_stopped:#if the voltage is over 4 then trigger an alert.
                        Alert_label.config(text="Over 5V Alert!")
                        if is_stopped:
                            toggle_stop()  # Turn off the system if not already stopped
                    elif turbine_voltage_value < 5 and is_stopped:# if the voltage is under 5 and the on/off button is pressed then chaneg the label to on and green
                        Alert_label.config( text="No Alerts", fg="green", font=("Arial", 20))# no alert
                        

                # Check if the line contains turbine current data
                match_current = re.search(r"Turbine Current:\s*(\d+)", line)
                if match_current:
                    turbine_current_value = int(match_current.group(1))  # Extract turbine current value
                    update_label(turbine_current_value_label, turbine_current_value)

                # Check if the line contains power data
                match_power = re.search(r"Power:\s*(\d+)", line)
                if match_power:
                    power_value = int(match_power.group(1))  # Extract power value
                    update_label(power_value_label, power_value)
                    

                    
                
                if "spinning" in line:# if it findes the word spinning in the line then 
                    turbine_status_label.config(text="Turbine is spinning", fg="green")  # Update label to show "spinning" in green
                elif "not " in line: # if it findes the word not in the line then 
                    turbine_status_label.config(text="Not spinning", fg="red")  # Update label to show "not spinning" in red

                
                    
                

            except Exception as e:# if it cant read the data then
                print("Error reading serial data:", e)# it will print" Error reading serial data:"

        time.sleep(0.01)

def update_total_power():#updates the total power generated every 0.1 second
    global power_value, total_power
    total_power += power_value
    Power = total_power/10
    update_label(Total_power_value_label, Power)
    window.after(100, update_total_power) 




def update_label(label, value):
    # This function is to ensure updates happen in the main Tkinter thread
    label.config(text=value)





#-------------------------------------------------------------------------------------------------------------------------------------------------
#                                                                   Left Panel
turbine_status_label = tk.Label(control_frame, text="Turbine Status", font=("Arial", 14))
turbine_status_label.grid(row=2, column=0, columnspan=5, sticky="n")
# Create main screen widgets (for logged-in users)
Turbine_label = tk.Label(control_frame, text="Wind Turbine", font=("Arial", 24))
Turbine_label.grid(row=0, column=0, columnspan=5, sticky="n")

# Stop button
stop_button = tk.Button(control_frame, text="Off", font=("Arial", 14), bg="red", command=toggle_stop)
stop_button.grid(row=1, column=0, columnspan=5, sticky="n")

# Angle display label
angle_label = tk.Label(control_frame, text=f"Angle: {servo_angle}°", font=("Arial", 20))
angle_label.grid(row=6, column=0, columnspan=5, sticky="n")



automatic_button = tk.Button(control_frame, text="Automatic", font=("Arial", 14), bg="red", command=toggle_automatic_mode)
automatic_button.grid(row=9, column=2, columnspan=2, sticky="n")

manual_button = tk.Button(control_frame, text="Manual", font=("Arial", 14), bg="green", command=toggle_manual_mode)
manual_button.grid(row=9, column=0, columnspan=2, sticky="n")

# Increase Angle button above the up arrow
up_button = tk.Button(control_frame, text="Increase Angle", font=("Arial", 14), command=increase_angle)
up_button.grid(row=4, column=0, columnspan=5, sticky="n")

# Up arrow label
up_arrow_label = tk.Label(control_frame, text="↑", font=("Arial", 24))
up_arrow_label.grid(row=5, column=0, columnspan=5, sticky="n")

# Down arrow label
down_arrow_label = tk.Label(control_frame, text="↓", font=("Arial", 24))
down_arrow_label.grid(row=7, column=0, columnspan=5, sticky="n")

# Decrease Angle button below the down arrow
down_button = tk.Button(control_frame, text="Decrease Angle", font=("Arial", 14), command=decrease_angle)
down_button.grid(row=8, column=0, columnspan=5, sticky="n")

#-------------------------------------------------------------------------------------------------------------------------------------------------
#                                                                   Right Panel
Turbine_data = tk.Label(control_frame, text="Wind Turbine data", font=("Arial", 25))
Turbine_data.grid(row=0, column=6,columnspan=4, sticky="n")

#-------------------------------------------------------------------------------------------------------------------------------------------------
#                                                                       temp
# Outside Temperature labels (description and value)
outside_temp_label = tk.Label(control_frame, text="Outside Temperature(C°):", font=("Arial", 14))
outside_temp_label.grid(row=1, column=6,columnspan=1, sticky="n")


outside_temp_value_label = tk.Label(control_frame, text="0", font=("Arial", 14))
outside_temp_value_label.grid(row=2, column=6,columnspan=1, sticky="n")

# Turbine Temperature labels (description and value)
turbine_temp_label = tk.Label(control_frame, text="Turbine Temperature(C°):", font=("Arial", 14))
turbine_temp_label.grid(row=1, column=8,columnspan=1, sticky="n")

turbine_temp_value_label = tk.Label(control_frame, text="0", font=("Arial", 14))
turbine_temp_value_label.grid(row=2, column=8,columnspan=1, sticky="n")

turbine_voltage_label = tk.Label(control_frame, text="Turbine Voltage(V):", font=("Arial", 14))
turbine_voltage_label.grid(row=3, column=6,columnspan=1, sticky="n")
turbine_voltage_value_label = tk.Label(control_frame, text="0", font=("Arial", 14))
turbine_voltage_value_label.grid(row=4, column=6,columnspan=1, sticky="n")

turbine_current_label = tk.Label(control_frame, text="Turbine Current(mA):", font=("Arial", 14))
turbine_current_label.grid(row=3, column=8,columnspan=1, sticky="n")
turbine_current_value_label = tk.Label(control_frame, text="0", font=("Arial", 14))
turbine_current_value_label.grid(row=4, column=8,columnspan=1, sticky="n")

power_label = tk.Label(control_frame, text="Powerm(W):", font=("Arial", 14))
power_label.grid(row=5, column=6,columnspan=1, sticky="n")
power_value_label = tk.Label(control_frame, text="0", font=("Arial", 14))
power_value_label.grid(row=6, column=6,columnspan=1, sticky="n")

total_power_label = tk.Label(control_frame, text="Total Power(mW):", font=("Arial", 14))
total_power_label.grid(row=5, column=8,columnspan=1, sticky="n")
Total_power_value_label = tk.Label(control_frame, text="0", font=("Arial", 14))
Total_power_value_label.grid(row=6, column=8,columnspan=1, sticky="n")

RPM_label = tk.Label(control_frame, text="Turbine RPM:", font=("Arial", 14))
RPM_label.grid(row=7, column=7,columnspan=1, sticky="n")
RPM_value_label = tk.Label(control_frame, text="0", font=("Arial", 14))
RPM_value_label.grid(row=8, column=7,columnspan=1, sticky="n")


#-------------------------------------------------------------------------------------------------------------------------------------------------
#                                               Aletr 
Alert_label = tk.Label(control_frame, text="No Alerts", fg="green", font=("Arial", 20))
Alert_label.grid(row=3, column=0,columnspan=5, sticky="n")


#-------------------------------------------------------------------------------------------------------------------------------------------------

#creates a grid in the tkinter winder so you can presisly place your labels and buttons
control_frame.grid_rowconfigure(0, weight=1)  # First row
control_frame.grid_rowconfigure(1, weight=1)  
control_frame.grid_rowconfigure(2, weight=1)  
control_frame.grid_rowconfigure(3, weight=2)  
control_frame.grid_rowconfigure(4, weight=1)  
control_frame.grid_rowconfigure(5, weight=1)
control_frame.grid_rowconfigure(6, weight=1)
control_frame.grid_rowconfigure(7, weight=1)
control_frame.grid_rowconfigure(8, weight=1)
control_frame.grid_rowconfigure(9, weight=1)
control_frame.grid_rowconfigure(10, weight=1)

control_frame.grid_columnconfigure(0, weight=1)  # first column
control_frame.grid_columnconfigure(1, weight=1)
control_frame.grid_columnconfigure(2, weight=1)
control_frame.grid_columnconfigure(3, weight=1)
control_frame.grid_columnconfigure(4, weight=1)
control_frame.grid_columnconfigure(5, weight=1)
control_frame.grid_columnconfigure(6, weight=1)
control_frame.grid_columnconfigure(7, weight=1)
control_frame.grid_columnconfigure(8, weight=1)
control_frame.grid_columnconfigure(9, weight=1)
control_frame.grid_columnconfigure(10, weight=1)






wb = openpyxl.Workbook() # Creates a new Excel workbook object.
ws = wb.active#gets the worksheet
ws.title = "Data"#names the worksheets data

# Writes the headersin Excel file
ws.append(["Time(s)", "Power(mw)","total_power(mw)", "Wind Direction(°)"])

# Function to update the seconds,power,totalpower, and servo angle in the excel file every second
def update_sum():
    global  seconds_counter, power_value, servo_angle, total_power

    while True:
        time.sleep(1)
        total_power += power_value # total power = the sum of all the power values after every second
        
        
        Total_power_value_label.config(text=f"{total_power}")  # Update total power label
        

        # Log values to the Excel file
        ws.append([seconds_counter, power_value, total_power, servo_angle])
        print(f"{seconds_counter}, {power_value}, {total_power}, {servo_angle}")
        # Update charts after logging data
        
        update_charts()
        for row in range(2, ws.max_row + 1):  # Start from row 2 to avoid the header
            cell = ws.cell(row=row, column=2)  # Column 2 is the "Power" column
            cell.number_format = '#,##0.00'  # Format as number with two decimal places

        seconds_counter += 1  # Increment the time counter

def update_charts():
    # Create Power Over Time chart
    chart = LineChart()
    chart.title = "Power Over Time"
    chart.style = 12
    chart.x_axis.title = "Time (s)"
    chart.y_axis.title = "Power (mw)"
    chart.y_axis.scaling.min = 0
    chart.y_axis.scaling.max = 100
    chart.legend = None
    # Create total Power Over Time chart
    chart3 = LineChart()
    chart3.title = "Total Power Produced"
    chart3.style = 12
    chart3.x_axis.title = "Time (s)"
    chart3.y_axis.title = "Total Power (mw)"
    chart3.y_axis.scaling.min = 0
    chart3.y_axis.scaling.max = total_power
    chart3.legend = None

    # Create Wind Direction chart over time
    chart2 = LineChart()
    chart2.title = "Wind Direction"
    chart2.style = 13
    chart2.x_axis.title = "Time (s)"
    chart2.y_axis.title = "Wind Direction (°)"
    chart2.y_axis.scaling.min = 0
    chart2.y_axis.scaling.max = 180
    chart2.legend = None

    # Reference for data in the charts
    power_data = Reference(ws, min_col=2, min_row=1, max_col=2, max_row=seconds_counter)
    chart.add_data(power_data, titles_from_data=True)

    Total_power_data = Reference(ws, min_col=3, min_row=1, max_col=3, max_row=seconds_counter)
    chart3.add_data(Total_power_data, titles_from_data=True)

    wind_direction_data = Reference(ws, min_col=4, min_row=1, max_col=4, max_row=seconds_counter)
    chart2.add_data(wind_direction_data, titles_from_data=True)

    # Add the charts to the sheet
    ws.add_chart(chart, "E1")  # Power chart at E1
    ws.add_chart(chart3, "N1")  # Power chart at E1
    ws.add_chart(chart2, "E20")  # Wind direction chart at E20

    # Save the workbook with the new data and charts
    wb.save(f"C:\\Users\\Liam\\OneDrive - Algonquin College\\Desktop\\(2)Supervisory Control and Data Acquisition\\Independant lab\\Python\\Logs\\logs_{current_datetime}.xlsx")
        
# theading to make sure that the functions can exicute without interfierign with eachother
serial_thread = threading.Thread(target=read_serial_data, daemon=True)
logging_thread = threading.Thread(target=update_sum, daemon=True)
power_thread = threading.Thread(target=update_total_power, daemon=True)
power_thread.start()
logging_thread.start()
serial_thread.start()



# Start the Tkinter event loop
window.mainloop()
























